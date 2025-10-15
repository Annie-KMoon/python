from openpyxl import Workbook

wb = Workbook()

ws = wb.active #워크북내 시트 선택
ws.title = '기본시트' #타이틀 설정

ws1 = wb.create_sheet() #새로운시트 생성 #active(기존시트 선택)
ws1.title = '나의시트'
ws1.sheet_properties.tabColor = 'FF00DD'

ws2 = wb.create_sheet('너의시트')

your_sheet = wb['너의시트']
your_sheet['A1'] = '테스트' #값입력(셀주소)

copy_sheet = wb.copy_worksheet(your_sheet) #시트복사
copy_sheet.title = '카피시트'

print(wb.sheetnames)
wb.save('data/sample.xlsx')
wb.close()