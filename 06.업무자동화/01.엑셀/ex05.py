from openpyxl import load_workbook

#엑셀파일 읽기
wb = load_workbook('data/sample.xlsx')
ws = wb.active

for row in ws.iter_rows(min_row=2, max_row=5): #첫행외 값행만 출력-min_row ~행까지 출력 max_row
    for col in row:
        print(col.value, end=',')
    print()

